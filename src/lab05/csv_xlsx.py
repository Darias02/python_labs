# CSV→XLSX
# B
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import csv
from pathlib import Path


def column_width(worksheet):  # автоматическая ширина колонок
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        # максимальная длина текста в колонке
        for cell in column:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = max(max_length + 2, 8)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    csv_file = Path(csv_path)
    if not csv_file.exists():
        raise FileNotFoundError(f"CSV файл не найден: {csv_path}")
    if csv_file.suffix.lower() != ".csv":
        raise ValueError(f"Файл {csv_file} не csv")
    xlsx_file = Path(xlsx_path)
    if xlsx_file.suffix.lower() != ".xlsx":
        raise ValueError(f"Файл {xlsx_file} не xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    try:
        with csv_file.open("r", encoding="utf-8", newline="") as f:
            csv_reader = csv.reader(f)

            for row_idx, row in enumerate(csv_reader, 1):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        column_width(ws)
        xlsx_file = Path(xlsx_path)
        xlsx_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(xlsx_path)
        print(f"конвертировано CSV → XLSX: {csv_path} → {xlsx_path}")
        print(f"размер таблицы: {ws.max_row} × {ws.max_column}")

    except Exception as e:
        raise ValueError(f"Ошибка конвертации CSV -> XLSX: {str(e)}")


if __name__ == "__main__":
    try:
        csv_to_xlsx("data/samples/people.csv", "data/out/people.xlsx")
        print("Тест CSV → XLSX")

        csv_to_xlsx("data/samples/cities.csv", "data/out/cities.xlsx")
        print("Тест cities.csv → XLSX")

    except Exception as e:
        print(f"ошибка: {e}")
